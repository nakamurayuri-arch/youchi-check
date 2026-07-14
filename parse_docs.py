# -*- coding: utf-8 -*-
"""解析結果dict → 評価表.xlsx（サマリ＋生データ）＋ チェック結果.md を生成。"""
from __future__ import annotations
import os, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pipeline

NAVY, TEAL, RED, GREY = "20376C", "2E6D8A", "CB0F4B", "F1EFE8"
FONT = "Meiryo UI"
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def _c(ws, coord, val, bold=False, color="000000", fill=None, size=10, align="left"):
    cell = ws[coord]; cell.value = val
    cell.font = Font(name=FONT, bold=bold, color=color, size=size)
    if fill: cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(vertical="center", horizontal=align, wrap_text=True)
    cell.border = BORDER


def _g(d, *keys, default="要確認"):
    cur = d
    for k in keys:
        if isinstance(cur, dict) and k in cur and cur[k] not in (None, "", []):
            cur = cur[k]
        else:
            return default
    return cur


def _flatten(prefix, obj, rows):
    if isinstance(obj, dict):
        for k, v in obj.items():
            _flatten(f"{prefix}.{k}" if prefix else str(k), v, rows)
    elif isinstance(obj, list):
        rows.append((prefix, json.dumps(obj, ensure_ascii=False)))
    else:
        rows.append((prefix, obj))


def build(result: dict, meta: dict, outdir: str):
    os.makedirs(outdir, exist_ok=True)
    name = meta.get("案件名", "案件")

    wb = Workbook()

    # ---- Sheet1: サマリ ----
    ws = wb.active; ws.title = "サマリ"; ws.sheet_view.showGridLines = False
    for i, w in enumerate([26, 60], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    _c(ws, "A1", f"{name} ｜ 案件チェック サマリ", bold=True, color=NAVY, size=14)
    ws.merge_cells("A1:B1")

    land = result.get("土地_登記公図", {})
    elec = result.get("電力申請", {})
    tl   = result.get("タイムライン", {})
    loc  = result.get("立地", {})
    haz  = _g(loc, "ハザード", default={})
    zone = _g(loc, "区域判定", default={})
    farm = _g(loc, "農地ステータス", default={})

    def haz_hit(k):
        v = haz.get(k) if isinstance(haz, dict) else None
        if isinstance(v, dict):
            return "該当" if v.get("該当") else ("非該当" if v.get("該当") is False else "要確認")
        return "要確認"

    summary = [
        ("所在", meta.get("所在", "要確認")),
        ("緯度経度", meta.get("緯度経度", _g(result, "入力", "lat"))),
        ("── 案件化・電力 ──", ""),
        ("案件化", _g(tl, "案件化")),
        ("受電日(推定)", _g(tl, "受電日_推定")),
        ("運開JEPX/EPRX(推定)", f"{_g(tl,'運開JEPX_推定')} / {_g(tl,'運開EPRX_推定')}"),
        ("受付番号", _g(elec, "受付番号")),
        ("接続検討回答日", _g(elec, "接続検討回答日")),
        ("受電電力kW", _g(elec, "受電電力kW")),
        ("保証金支払", json.dumps(_g(tl, "保証金支払_推定", default="要確認"), ensure_ascii=False)),
        ("負担金支払", json.dumps(_g(tl, "負担金支払_推定", default="要確認"), ensure_ascii=False)),
        ("── 権利（登記・公図）──", ""),
        ("地番 / 地目", f"{_g(land,'地番')} / {_g(land,'地目_現況推定')}"),
        ("地積(候補㎡)", json.dumps(_g(land, "地積_候補㎡", default="要確認"), ensure_ascii=False)),
        ("現所有者(推定)", _g(land, "現所有者_推定")),
        ("甲区 最終更新", _g(land, "最終更新_推定")),
        ("乙区(抵当権)", _g(land, "乙区記載")),
        ("公図種別", _g(land, "公図種別")),
        ("── 立地（自動判定）──", ""),
        ("洪水浸水想定", haz_hit("洪水浸水想定(想定最大規模)")),
        ("土砂(土石流/急傾斜/地すべり)",
         f"{haz_hit('土砂災害警戒区域(土石流)')}/{haz_hit('土砂災害警戒区域(急傾斜)')}/{haz_hit('土砂災害警戒区域(地すべり)')}"),
        ("津波 / 高潮", f"{haz_hit('津波浸水想定')} / {haz_hit('高潮浸水想定')}"),
        ("標高/傾斜", json.dumps(_g(loc, "標高傾斜", default="要確認"), ensure_ascii=False)),
        ("最寄り道路", json.dumps(_g(loc, "最寄り道路", default="要確認"), ensure_ascii=False)),
        ("最寄り建物m(近隣住宅目安)", _g(loc, "最寄り建物m")),
        ("海岸距離m(重塩害)", _g(loc, "海岸距離m")),
        ("農地(第何種 推定)", _g(farm, "推定結果", "第何種_推定")),
    ]
    r = 3
    for k, v in summary:
        header = str(k).startswith("──")
        _c(ws, f"A{r}", k, bold=True, color=("FFFFFF" if header else NAVY),
           fill=(TEAL if header else GREY), size=10)
        if header:
            _c(ws, f"B{r}", "", fill=TEAL)
        else:
            val = "要確認" if v in (None, "", "要確認") else v
            col = RED if val == "要確認" else "000000"
            _c(ws, f"B{r}", val, color=col, bold=(val == "要確認"))
        ws.row_dimensions[r].height = 22
        r += 1

    # 残タスク（常に赤字）
    r += 1
    _c(ws, f"A{r}", "自動化不可＝人が確認", bold=True, color="FFFFFF", fill=RED); _c(ws, f"B{r}", "", fill=RED)
    r += 1
    for item in pipeline.RESIDUAL:
        _c(ws, f"A{r}", "要確認", bold=True, color=RED, align="center")
        _c(ws, f"B{r}", item, color=RED)
        ws.row_dimensions[r].height = 20
        r += 1

    # ---- Sheet2: 生データ ----
    ws2 = wb.create_sheet("生データ"); ws2.sheet_view.showGridLines = False
    for i, w in enumerate([48, 60], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    _c(ws2, "A1", "解析 生データ（トレーサビリティ用）", bold=True, color=NAVY, size=12)
    ws2.merge_cells("A1:B1")
    rows = []; _flatten("", result, rows)
    rr = 3
    for k, v in rows:
        _c(ws2, f"A{rr}", k, size=9, color=NAVY)
        _c(ws2, f"B{rr}", "" if v is None else str(v), size=9)
        rr += 1

    xlsx = os.path.join(outdir, f"{name}_チェック結果.xlsx")
    wb.save(xlsx)

    # ---- md ----
    md = os.path.join(outdir, f"{name}_チェック結果.md")
    with open(md, "w", encoding="utf-8") as f:
        f.write(f"# {name} 案件チェック結果\n\n")
        f.write(f"所在: {meta.get('所在','要確認')} / 緯度経度: {meta.get('緯度経度','')}\n\n")
        f.write("## サマリ\n\n| 項目 | 値 |\n|---|---|\n")
        for k, v in summary:
            if str(k).startswith("──"):
                f.write(f"| **{k.strip('─ ')}** | |\n")
            else:
                f.write(f"| {k} | {'要確認' if v in (None,'','要確認') else v} |\n")
        f.write("\n## 自動化不可（人が確認）\n\n")
        for item in pipeline.RESIDUAL:
            f.write(f"- [ ] {item}\n")

    # ---- raw json ----
    js = os.path.join(outdir, f"{name}_results.json")
    with open(js, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    return {"xlsx": xlsx, "md": md, "json": js}
